from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime


class ExcelService:
    """Service for generating reports in Excel"""
    
    # Theme colors
    COLOR_HEADER = 'D97706'  # Amber-600
    COLOR_ALERT = 'DC2626'   # Red-600
    COLOR_LIGHT = 'FEF3C7'   # Amber-100
    
    @staticmethod
    def _format_currency(value):
        """Formats monetary values"""
        try:
            return float(value)
        except:
            return 0.0

    @staticmethod
    def _format_date(date):
        """Formats dates"""
        if date:
            return date.strftime('%d/%m/%Y %H:%M')
        return "N/A"

    @staticmethod
    def _apply_header_style(ws, row=1):
        """Applies style to the header"""
        header_fill = PatternFill(start_color=ExcelService.COLOR_HEADER, 
                                  end_color=ExcelService.COLOR_HEADER, 
                                  fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        for cell in ws[row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

    @staticmethod
    def _auto_adjust_columns(ws):
        """Automatically adjusts column widths"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    @staticmethod
    def _add_borders(ws, start_row=1, end_row=None):
        """Adds borders to cells"""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        if end_row is None:
            end_row = ws.max_row
        
        for row in ws.iter_rows(min_row=start_row, max_row=end_row):
            for cell in row:
                cell.border = thin_border

    @staticmethod
    def generate_orders_report_excel(pedidos, filtros):
        """Generates orders report in Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Orders Report"
        
        # Title
        ws['A1'] = 'ORDERS REPORT'
        ws['A1'].font = Font(bold=True, size=16, color=ExcelService.COLOR_HEADER)
        ws.merge_cells('A1:E1')
        
        # Generation date
        ws['A2'] = f'Generated: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        ws['A2'].font = Font(italic=True, size=9)
        ws.merge_cells('A2:E2')
        
        # Filters
        row = 3
        if filtros:
            filtros_text = "Filters: "
            if filtros.get('fecha_desde'):
                filtros_text += f"From {filtros['fecha_desde']} "
            if filtros.get('fecha_hasta'):
                filtros_text += f"To {filtros['fecha_hasta']} "
            if filtros.get('estado'):
                filtros_text += f"Status: {filtros['estado']} "
            
            ws[f'A{row}'] = filtros_text
            ws[f'A{row}'].font = Font(italic=True, size=9)
            ws.merge_cells(f'A{row}:E{row}')
            row += 1
        
        # Summary
        row += 1
        total_pedidos = len(pedidos)
        total_ingresos = sum(p.total for p in pedidos)
        
        ws[f'A{row}'] = 'Total Orders:'
        ws[f'B{row}'] = total_pedidos
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = 'Total Income:'
        ws[f'B{row}'] = ExcelService._format_currency(total_ingresos)
        ws[f'B{row}'].number_format = '$#,##0.00'
        ws[f'A{row}'].font = Font(bold=True)
        
        # Space
        row += 2
        
        # Headers
        headers = ['ID', 'Customer', 'Date', 'Status', 'Total']
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
        
        ExcelService._apply_header_style(ws, row)
        
        # Data
        for pedido in pedidos:
            row += 1
            ws.cell(row=row, column=1, value=pedido.id)
            ws.cell(row=row, column=2, value=pedido.usuario.get_full_name() or pedido.usuario.username)
            ws.cell(row=row, column=3, value=ExcelService._format_date(pedido.fecha_pedido))
            ws.cell(row=row, column=4, value=pedido.get_estado_display())
            ws.cell(row=row, column=5, value=ExcelService._format_currency(pedido.total))
            ws.cell(row=row, column=5).number_format = '$#,##0.00'
        
        # Final adjustments
        ExcelService._auto_adjust_columns(ws)
        ExcelService._add_borders(ws, start_row=row - len(pedidos), end_row=row)
        
        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    def generate_sales_report_excel(datos_ventas, filtros):
        """Generates sales report in Excel"""
        wb = Workbook()
        
        # Sheet 1: Summary
        ws1 = wb.active
        ws1.title = "Sales Summary"
        
        # Title
        ws1['A1'] = 'SALES REPORT'
        ws1['A1'].font = Font(bold=True, size=16, color=ExcelService.COLOR_HEADER)
        ws1.merge_cells('A1:B1')
        
        # Period
        ws1['A2'] = f"Period: {filtros.get('fecha_desde', 'Start')} - {filtros.get('fecha_hasta', 'Today')}"
        ws1['A2'].font = Font(italic=True, size=9)
        ws1.merge_cells('A2:B2')
        
        # Metrics
        row = 4
        metrics = [
            ('Total Income', ExcelService._format_currency(datos_ventas['total_ingresos']), '$#,##0.00'),
            ('Orders Count', datos_ventas['cantidad_pedidos'], '0'),
            ('Average Ticket', ExcelService._format_currency(datos_ventas['ticket_promedio']), '$#,##0.00'),
        ]
        
        for metric_name, metric_value, number_format in metrics:
            ws1[f'A{row}'] = metric_name
            ws1[f'B{row}'] = metric_value
            ws1[f'A{row}'].font = Font(bold=True)
            ws1[f'B{row}'].number_format = number_format
            ws1[f'A{row}'].fill = PatternFill(start_color=ExcelService.COLOR_LIGHT, 
                                             end_color=ExcelService.COLOR_LIGHT, 
                                             fill_type='solid')
            ws1[f'B{row}'].fill = PatternFill(start_color=ExcelService.COLOR_LIGHT, 
                                             end_color=ExcelService.COLOR_LIGHT, 
                                             fill_type='solid')
            row += 1
        
        # Sheet 2: Top Dishes
        ws2 = wb.create_sheet("Top Dishes")
        
        ws2['A1'] = 'TOP 10 BEST SELLING DISHES'
        ws2['A1'].font = Font(bold=True, size=14, color=ExcelService.COLOR_HEADER)
        ws2.merge_cells('A1:D1')
        
        # Headers
        headers = ['#', 'Dish', 'Quantity', 'Income']
        for col, header in enumerate(headers, 1):
            ws2.cell(row=3, column=col, value=header)
        
        ExcelService._apply_header_style(ws2, 3)
        
        # Data
        row = 4
        for idx, plato in enumerate(datos_ventas['platos_mas_vendidos'][:10], 1):
            ws2.cell(row=row, column=1, value=idx)
            ws2.cell(row=row, column=2, value=plato['nombre'])
            ws2.cell(row=row, column=3, value=plato['cantidad'])
            ws2.cell(row=row, column=4, value=ExcelService._format_currency(plato['ingresos']))
            ws2.cell(row=row, column=4).number_format = '$#,##0.00'
            row += 1
        
        # Adjustments
        ExcelService._auto_adjust_columns(ws1)
        ExcelService._auto_adjust_columns(ws2)
        ExcelService._add_borders(ws2, start_row=3, end_row=row-1)
        
        # Save
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    def generate_inventory_report_excel(productos, filtros):
        """Generates inventory report in Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory Report"
        
        # Title
        ws['A1'] = 'INVENTORY REPORT'
        ws['A1'].font = Font(bold=True, size=16, color=ExcelService.COLOR_HEADER)
        ws.merge_cells('A1:E1')
        
        # Date
        ws['A2'] = f'Generated: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        ws['A2'].font = Font(italic=True, size=9)
        ws.merge_cells('A2:E2')
        
        # Summary
        row = 4
        total_productos = len(productos)
        productos_bajo_stock = sum(1 for p in productos if p.stock_actual <= p.alerta_stock)
        valor_total = sum(p.stock_actual * p.precio for p in productos)
        
        ws[f'A{row}'] = 'Total Products:'
        ws[f'B{row}'] = total_productos
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = 'Low Stock Products:'
        ws[f'B{row}'] = productos_bajo_stock
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = 'Total Inventory Value:'
        ws[f'B{row}'] = ExcelService._format_currency(valor_total)
        ws[f'B{row}'].number_format = '$#,##0.00'
        ws[f'A{row}'].font = Font(bold=True)
        
        # Space
        row += 2
        
        # Headers
        headers = ['Product', 'Current Stock', 'Alert', 'Price', 'Total Value']
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
        
        ExcelService._apply_header_style(ws, row)
        
        # Data
        alert_font = Font(bold=True, color=ExcelService.COLOR_ALERT)
        alert_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
        
        for producto in productos:
            row += 1
            bajo_stock = producto.stock_actual <= producto.alerta_stock
            
            ws.cell(row=row, column=1, value=producto.nombre)
            ws.cell(row=row, column=2, value=producto.stock_actual)
            ws.cell(row=row, column=3, value=producto.alerta_stock)
            ws.cell(row=row, column=4, value=ExcelService._format_currency(producto.precio))
            ws.cell(row=row, column=4).number_format = '$#,##0.00'
            ws.cell(row=row, column=5, value=ExcelService._format_currency(producto.stock_actual * producto.precio))
            ws.cell(row=row, column=5).number_format = '$#,##0.00'
            
            # Highlight low stock products
            if bajo_stock:
                for col in range(1, 6):
                    ws.cell(row=row, column=col).font = alert_font
                    ws.cell(row=row, column=col).fill = alert_fill
        
        # Adjustments
        ExcelService._auto_adjust_columns(ws)
        ExcelService._add_borders(ws, start_row=row - len(productos), end_row=row)
        
        # Save
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    def generate_reservations_report_excel(reservas, filtros):
        """Generates reservations report in Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Reservations Report"
        
        # Title
        ws['A1'] = 'RESERVATIONS REPORT'
        ws['A1'].font = Font(bold=True, size=16, color=ExcelService.COLOR_HEADER)
        ws.merge_cells('A1:E1')
        
        # Date
        ws['A2'] = f'Generated: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        ws['A2'].font = Font(italic=True, size=9)
        ws.merge_cells('A2:E2')
        
        # Filters
        row = 3
        if filtros:
            filtros_text = "Filters: "
            if filtros.get('fecha_desde'):
                filtros_text += f"From {filtros['fecha_desde']} "
            if filtros.get('fecha_hasta'):
                filtros_text += f"To {filtros['fecha_hasta']} "
            if filtros.get('estado'):
                filtros_text += f"Status: {filtros['estado']} "
            
            ws[f'A{row}'] = filtros_text
            ws[f'A{row}'].font = Font(italic=True, size=9)
            ws.merge_cells(f'A{row}:E{row}')
            row += 1
        
        # Summary
        row += 1
        total_reservas = len(reservas)
        total_personas = sum(r.cantidad_personas for r in reservas)
        
        ws[f'A{row}'] = 'Total Reservations:'
        ws[f'B{row}'] = total_reservas
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = 'Total People:'
        ws[f'B{row}'] = total_personas
        ws[f'A{row}'].font = Font(bold=True)
        
        # Space
        row += 2
        
        # Headers
        headers = ['ID', 'Customer', 'Date', 'People', 'Status']
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
        
        ExcelService._apply_header_style(ws, row)
        
        # Data
        for reserva in reservas:
            row += 1
            ws.cell(row=row, column=1, value=reserva.id)
            ws.cell(row=row, column=2, value=reserva.usuario.get_full_name() or reserva.usuario.username)
            ws.cell(row=row, column=3, value=ExcelService._format_date(reserva.fecha_reserva))
            ws.cell(row=row, column=4, value=reserva.cantidad_personas)
            ws.cell(row=row, column=5, value=reserva.get_estado_display())
        
        # Adjustments
        ExcelService._auto_adjust_columns(ws)
        ExcelService._add_borders(ws, start_row=row - len(reservas), end_row=row)
        
        # Save
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
