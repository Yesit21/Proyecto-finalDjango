from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO

class ExcelService:
    @staticmethod
    def export_orders(pedidos):
        wb = Workbook()
        ws = wb.active
        ws.title = "Pedidos"
        
        headers = ['ID', 'Cliente', 'Fecha', 'Estado', 'Total']
        ws.append(headers)
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        for pedido in pedidos:
            ws.append([
                pedido.id,
                pedido.cliente.username,
                pedido.fecha_pedido.strftime('%Y-%m-%d'),
                pedido.get_estado_display(),
                float(pedido.total)
            ])
        
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
